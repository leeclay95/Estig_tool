#!/usr/bin/env python3
"""
estig_tool.py — STIG workbook / XML / report helper

Flags you can combine (order is fixed):

  -i, --init           • Init workbook from template
  -c, --clear          • Clear data rows in a workbook
  -u, --update         • Import Not_Reviewed V-keys from .cklb scans
  -m, --manualestig    • Alias for --update (legacy manfindexcel)
  -g, --generate       • Generate/refresh XML answer-files from workbook
  -r, --report         • Recursively scan a directory for .cklb → build report

Run without flags for an interactive menu.

Requires: pip install openpyxl pandas
"""

import os, json, glob, shutil, sys, datetime, argparse, xml.etree.ElementTree as ET
from pathlib import Path
from collections import Counter, defaultdict
import re

import pandas as pd
from openpyxl import load_workbook

# ────────────────────────── CONSTANTS ────────────────────────────────────────
ACTIVE_SHORTNAMES = [
    "ADDomain","ADForest","AdobeAcrobatProDCContinuous","AdobeReaderDCContinuous",
    "Apache24SvrUnix","Apache24SvrWin","Apache24SiteUnix","Apache24SiteWin",
    "ApacheTomcatAS","CiscoXERtrNDM","CiscoXESwtchL2S","CiscoXESwtchNDM","Chrome",
    "IIS10Server","IIS10Site","IE11","JBoss","DotNET4","MSAccess2016","MSDefender",
    "MSEdge","MSExcel2016","MSExchange2016EdgeTP","MSExchange2016MB","MSExchange2019Edge",
    "MSExchange2019MB","MSGroove2013","MSOffice365","MSOfficeSystem2016","MSOneDrive",
    "MSOneNote2013","MSOneNote2016","MSOutlook2013","MSOutlook2016","MSPowerPoint2016",
    "MSProject2016","MSPublisher2016","MSSPDesigner2013","MSSkype2016","SQL2016DB",
    "SQL2016Instance","MSVisio2016","MSWord2016","Firefox","Oracle7","Oracle8","RGSRKE2",
    "RHEL8","RHEL9","TrellixENS10xLocal","Ubuntu20","Ubuntu22","Win10","Win11",
    "WinFirewall","WinServer2016","WinServer2019","WinServer2022","WinServerDNS"
]

# we will extract timestamp from filename via this regex:
TIMESTAMP_RX = re.compile(r'(\d{8}-\d{6})\.cklb$', re.I)
# prefix in JSON "title" field:
TITLE_PREFIX = "Evaluate-STIG_"

# ────────────────────────── HELPERS ───────────────────────────────────────────
def ts_now(fmt="%Y-%m-%d %H:%M:%S") -> str:
    return datetime.datetime.now().strftime(fmt)

def prompt_path(msg: str, default: str | None = None) -> str:
    inp = input(f"{msg}{f' [{default}]' if default else ''}: ").strip().strip('"').strip("'")
    return os.path.expanduser(inp or (default or ""))

def yes(msg: str, default=True) -> bool:
    yn = ("Y/n","y/N")[not default]
    while True:
        a = input(f"{msg} ({yn}) ").strip().lower()
        if not a: return default
        if a in ("y","yes"): return True
        if a in ("n","no"):  return False
        print("Enter y or n.")

def pretty(n: int) -> str:
    return f"{n:,}"

# ═════════════════════════════════════════════════════════════════════════════
# 1. INIT WORKBOOK — create sheets & “AnswerKey Name” column
# ═════════════════════════════════════════════════════════════════════════════
def init_workbook():
    src = prompt_path("Template workbook path (.xlsx)")
    dst = Path(prompt_path("New workbook to create")).expanduser()
    if dst.exists() and not yes("Overwrite existing?", False):
        print("Aborted."); return
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(src, dst)
    wb = load_workbook(dst)
    headers = [c.value for c in wb["Chrome"][1]]
    for sn in ACTIVE_SHORTNAMES:
        if sn not in wb.sheetnames:
            ws = wb.create_sheet(sn)
            for i, h in enumerate(headers,1):
                ws.cell(1,i,h)
    for ws in wb.worksheets:
        if "AnswerKey Name" not in [c.value for c in ws[1]]:
            ws.cell(1, ws.max_column+1, "AnswerKey Name")
    wb.save(dst)
    print(f"✅ Created workbook at {dst}")

# ═════════════════════════════════════════════════════════════════════════════
# 2. CLEAR EXCEL — wipe rows below header
# ═════════════════════════════════════════════════════════════════════════════
def clear_excel():
    path = Path(prompt_path("Workbook to clear (.xlsx)")).expanduser()
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for c in row: c.value = None
    wb.save(path)
    print("✅ Cleared all data rows.")

# ═════════════════════════════════════════════════════════════════════════════
# 3. UPDATE EXCEL — import Not_Reviewed V-keys & optional XML
# ═════════════════════════════════════════════════════════════════════════════
def newest_cklb(root: str) -> dict[str,str]:
    """
    Walk root, open each .cklb JSON, extract sheet-shortname from its "title":
      "Evaluate-STIG_<ShortName>"
    and group by ShortName, keeping only the file with the highest
    timestamp (from the filename).
    """
    latest: dict[str,tuple[str,str]] = {}
    for dirpath, _, files in os.walk(root):
        for fn in files:
            if not fn.lower().endswith(".cklb"):
                continue
            m_ts = TIMESTAMP_RX.search(fn)
            if not m_ts:
                continue
            ts = m_ts.group(1)
            full = os.path.join(dirpath, fn)
            # load JSON to get title
            try:
                data = json.load(open(full, encoding="utf-8"))
            except Exception:
                continue
            title = data.get("title","")
            if not title.startswith(TITLE_PREFIX):
                continue
            short = title[len(TITLE_PREFIX):]
            # pick newest
            prev = latest.get(short)
            if not prev or ts > prev[0]:
                latest[short] = (ts, full)
    # return short → filepath
    return {short: path for short,(ts,path) in latest.items()}

def not_reviewed(path: str) -> list[str]:
    try:
        data = json.load(open(path, encoding="utf-8"))
    except Exception as e:
        print(f"⚠️ Skipping {path}: {e}")
        return []
    out = []
    for s in data.get("stigs", []):
        for r in s.get("rules", []):
            if str(r.get("status","")).lower() == "not_reviewed":
                vk = r.get("group_id") or r.get("vuln_id")
                if vk:
                    out.append(vk.strip())
    return out

def header_map(ws):
    hdrs = {str(c.value).strip().lower(): idx for idx,c in enumerate(ws[1],1) if c.value}
    def col(name: str):
        if name not in hdrs:
            idx = ws.max_column + 1
            ws.cell(1, idx, name)
            hdrs[name] = idx
        return hdrs[name]
    return {
        "vuln id":          col("vuln id"),
        "expectedstatus":   col("expectedstatus"),
        "validtruestatus":  col("validtruestatus"),
        "validtruecomment": col("validtruecomment"),
    }

def remove_old_comments(root: ET.Element):
    for e in list(root):
        if e.tag is ET.Comment and e.text and e.text.startswith("Script ran on"):
            root.remove(e)

def update_excel():
    wb_path  = Path(prompt_path("Workbook to update (.xlsx)")).expanduser()
    scan_dir = prompt_path("Directory containing .cklb files")
    vt_cmt   = input("ValidTrueComment (blank → STIG COMPLIANT): ").strip() or "STIG COMPLIANT"
    do_xml   = yes("Also update/create XML answer-files?", True)
    xml_dir  = Path(prompt_path("XML output directory")) if do_xml else None
    if xml_dir:
        xml_dir.mkdir(parents=True, exist_ok=True)

    stig_keys = {s: not_reviewed(p) for s,p in newest_cklb(scan_dir).items()}
    stig_keys = {s: ks for s,ks in stig_keys.items() if ks}

    wb = load_workbook(wb_path)
    for stig, keys in stig_keys.items():
        if stig not in wb.sheetnames:
            print(f"❌ Sheet '{stig}' missing – skipped.")
            continue
        ws = wb[stig]
        cols = header_map(ws)
        existing = {
            str(r[0]).strip()
            for r in ws.iter_rows(min_row=2,
                                  min_col=cols["vuln id"],
                                  max_col=cols["vuln id"],
                                  values_only=True)
            if r[0]
        }
        row = ws.max_row + 1
        for vk in keys:
            if vk in existing:
                continue
            ws.cell(row, cols["vuln id"], vk)
            ws.cell(row, cols["expectedstatus"], "Not_Reviewed")
            ws.cell(row, cols["validtruestatus"], "NotAFinding")
            ws.cell(row, cols["validtruecomment"], vt_cmt)
            row += 1
    wb.save(wb_path)
    print("✔️ Excel updated.")

    if not do_xml:
        return

    for stig, keys in stig_keys.items():
        xml = xml_dir / f"{stig}.xml"
        if xml.exists():
            tree = ET.parse(xml); root = tree.getroot()
        else:
            root = ET.Element("STIGComments", Name=stig)
            tree = ET.ElementTree(root)

        existing = {v.get("ID") for v in root.findall("Vuln")}
        added = []
        for vk in keys:
            if vk in existing:
                continue
            vul = ET.SubElement(root, "Vuln", ID=vk)
            ak  = ET.SubElement(vul, "AnswerKey", Name="DEFAULT")
            ET.SubElement(ak, "ExpectedStatus").text   = "Not_Reviewed"
            ET.SubElement(ak, "ValidationCode")
            ET.SubElement(ak, "ValidTrueStatus").text  = "NotAFinding"
            ET.SubElement(ak, "ValidTrueComment").text = vt_cmt
            ET.SubElement(ak, "ValidFalseStatus")
            ET.SubElement(ak, "ValidFalseComment")
            added.append(vk)

        if added:
            remove_old_comments(root)
            all_keys = sorted({v.get("ID") for v in root.findall("Vuln")})
            comment  = f"Script ran on {ts_now()} – Added V-keys: {', '.join(added)}"
            root.append(ET.Comment(comment))
            tree.write(xml, encoding="utf-8", xml_declaration=True)
            print(f"📝 {stig}: Added V-keys → {', '.join(added)}")

# ═════════════════════════════════════════════════════════════════════════════
# 4. GENERATE XML FROM WORKBOOK — export rows as XML answer-files
# ═════════════════════════════════════════════════════════════════════════════
def generate_xml():
    wb_file = Path(prompt_path("Workbook path (.xlsx)")).expanduser()
    out_dir = Path(prompt_path("XML output directory")).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)

    xls = pd.ExcelFile(wb_file)
    for sheet in xls.sheet_names:
        df = xls.parse(sheet)
        xml = out_dir / f"{sheet}.xml"
        if xml.exists():
            tree = ET.parse(xml); root = tree.getroot()
        else:
            root = ET.Element("STIGComments", Name=sheet)
            tree = ET.ElementTree(root)

        added = []
        for _, row in df.iterrows():
            vid = row.get("Vuln ID", row.get("V Key", ""))
            vid = "" if pd.isna(vid) else str(vid).strip()
            if not vid:
                continue

            key = row.get("AnswerKey Name", "")
            key = "" if pd.isna(key) else str(key).strip()
            if not key:
                key = "DEFAULT"

            exp = row.get("ExpectedStatus", "")
            exp = "" if pd.isna(exp) else str(exp).strip() or "Not_Reviewed"
            vts = row.get("ValidTrueStatus", "")
            vts = "" if pd.isna(vts) else str(vts).strip() or "NotAFinding"
            vtc = row.get("ValidTrueComment", "")
            vtc = "" if pd.isna(vtc) else str(vtc).strip()

            vuln = next((v for v in root.findall("Vuln") if v.get("ID")==vid), None)
            if vuln is None:
                vuln = ET.SubElement(root, "Vuln", ID=vid)

            if any(ak.get("Name")==key for ak in vuln.findall("AnswerKey")):
                continue

            ak = ET.SubElement(vuln, "AnswerKey", Name=key)
            ET.SubElement(ak, "ExpectedStatus").text   = exp
            ET.SubElement(ak, "ValidationCode")
            ET.SubElement(ak, "ValidTrueStatus").text  = vts
            ET.SubElement(ak, "ValidTrueComment").text = vtc
            ET.SubElement(ak, "ValidFalseStatus")
            ET.SubElement(ak, "ValidFalseComment")

            added.append(f"{vid}({key})")

        if added:
            comment = f"Script ran on {ts_now()} – Added: {', '.join(added)}"
            root.append(ET.Comment(comment))
            tree.write(xml, encoding="utf-8", xml_declaration=True)
            print(f"🛈 {sheet}: Added → {', '.join(added)}")
        else:
            print(f"✓ {sheet}: no new entries")

# ═════════════════════════════════════════════════════════════════════════════
# 5. REPORT — recursive scan for .cklb → Markdown findings report
# ═════════════════════════════════════════════════════════════════════════════
def extract_cklb(path: str):
    try:
        data = json.load(open(path, encoding="utf-8"))
    except Exception as e:
        print(f"❌ {path}: {e}")
        return None, [], 0, Counter()

    host = (
        data.get("host_name")
        or (data.get("targets",[{}])[0].get("host_name") if isinstance(data.get("targets"), list) else None)
        or data.get("target_data",{}).get("host_name")
        or "Unknown Host"
    )

    findings = []
    total = 0
    counts = Counter()
    for s in data.get("stigs", []):
        nm = s.get("stig_name","Unknown STIG")
        sc = Counter(r.get("status","").strip().lower() for r in s.get("rules", []))
        counts.update(sc)
        n = sum(sc.values())
        total += n
        findings.append((nm, sc))

    return host, findings, total, counts

def report():
    root_dir = Path(prompt_path("Directory to scan recursively for .cklb")).expanduser()
    files = list(root_dir.rglob("*.cklb"))
    if not files:
        print("❌ No .cklb files found under that directory."); return

    out = Path(prompt_path("Output Markdown file or folder")).expanduser()
    if out.is_dir():
        out = out / f"stig_report_{datetime.datetime.now():%Y%m%d-%H%M%S}.md"

    summary = defaultdict(int)
    details = []
    grand = Counter()
    for fp in files:
        host, findings, tot, cnt = extract_cklb(str(fp))
        if host:
            summary[host] += tot
            details.append((fp, host, findings))
            grand.update(cnt)

    with open(out, "w", encoding="utf-8") as fh:
        fh.write("# STIG Open Findings Summary Report\n\n")
        fh.write("## Summary by Host\n\n| Host | Total Findings |\n|------|----------------|\n")
        for h,t in summary.items():
            fh.write(f"| {h} | {pretty(t)} |\n")

        fh.write("\n## Detailed Findings by File\n\n")
        for fp, host, findings in details:
            fh.write(f"### File: `{fp.name}`\n- Host: **{host}**\n")
            for nm, sc in findings:
                cnt = sum(sc.values())
                fh.write(f"  - STIG: *{nm}* — **{pretty(cnt)}** findings\n")
                for st,c in sorted(sc.items()):
                    fh.write(f"    - {st.replace('_',' ').title()}: {pretty(c)}\n")
            fh.write("\n")

        impl = grand.get("not_a_finding",0)
        op   = grand.get("open",0)
        tot  = impl + op
        pct  = (impl/tot*100) if tot else 0
        fh.write("## STIG Implementation Summary\n\n")
        fh.write(f"- Total Evaluated: **{pretty(tot)}**\n")
        fh.write(f"- Compliant (Not a Finding): **{pretty(impl)}**\n")
        fh.write(f"- Non-compliant (Open): **{pretty(op)}**\n")
        fh.write(f"\n**Overall Implementation: {pct:.2f}%**\n")
        fh.write(f"\n---\n_Report generated on {ts_now()}_\n")

    print(f"✅ Markdown report saved to {out}")

# ═════════════════════════════════════════════════════════════════════════════
# 6. INTERACTIVE MENU
# ═════════════════════════════════════════════════════════════════════════════
def menu():
    while True:
        print("\n1) Init workbook   2) Clear   3) Update   4) Generate XML   5) Report   0) Exit")
        c = input("> ").strip()
        if c == "0":
            break
        {
            "1": init_workbook,
            "2": clear_excel,
            "3": update_excel,
            "4": generate_xml,
            "5": report
        }.get(c, lambda: print("Invalid choice"))()

# ═════════════════════════════════════════════════════════════════════════════
# 7. CLI ENTRYPOINT
# ═════════════════════════════════════════════════════════════════════════════
EPILOG = """
EXAMPLES
  python estig_tool.py -u
  python estig_tool.py -c -u -g
  python estig_tool.py -r
"""

def main():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawTextHelpFormatter,
        description="All-in-one STIG Excel/XML/report helper",
        epilog=EPILOG
    )
    parser.add_argument("-i","--init",        action="store_true",
                        help="Create a workbook from a template.")
    parser.add_argument("-c","--clear",       action="store_true",
                        help="Clear all data rows in a workbook.")
    parser.add_argument("-u","--update",      action="store_true",
                        help="Import Not_Reviewed V-keys into the workbook and optionally XML.")
    parser.add_argument("-m","--manualestig", action="store_true",
                        help="Alias for --update (legacy).")
    parser.add_argument("-g","--generate",    action="store_true",
                        help="Generate/refresh XML answer-files from the workbook.")
    parser.add_argument("-r","--report",      action="store_true",
                        help="Recursively scan for .cklb and build Markdown report.")
    args = parser.parse_args()

    steps = []
    if args.init:
        steps.append(init_workbook)
    if args.clear:
        steps.append(clear_excel)
    if args.update or args.manualestig:
        steps.append(update_excel)
    if args.generate:
        steps.append(generate_xml)
    if args.report:
        steps.append(report)

    if not steps:
        menu()
    else:
        for fn in steps:
            try:
                fn()
            except KeyboardInterrupt:
                sys.exit("\nInterrupted by user.")
            except Exception as e:
                print(f"❌ Error: {e}")

if __name__ == "__main__":
    main()
