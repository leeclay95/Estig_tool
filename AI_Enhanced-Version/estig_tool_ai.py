#!/usr/bin/env python3
"""
estig_tool.py — STIG workbook / XML / report helper with AI PowerShell Generation

Flags you can combine (order is fixed):

  -i, --init           • Init workbook from template
  -c, --clear          • Clear data rows in a workbook
  -u, --update         • Import Not_Reviewed V-keys from .cklb scans
  -m, --manualestig    • Alias for --update (legacy manfindexcel)
  -g, --generate       • Generate/refresh XML answer-files from workbook
  -r, --report         • Recursively scan a directory for .cklb → build report
  -p, --powershell     • Generate PowerShell validation code using AI
  -a, --aiconfig       • Configure AI model settings
  -l, --library        • Browse STIG library and generate PowerShell from rules

Run without flags for an interactive menu.

Requires: pip install openpyxl pandas requests
"""

import os, json, glob, shutil, sys, datetime, argparse, xml.etree.ElementTree as ET
from pathlib import Path
from collections import Counter, defaultdict
import re
import zipfile
import requests
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

# ────────────────────────── CONSTANTS ────────────────────────────────────────
ACTIVE_SHORTNAMES = [
    "ADDomain","ADForest","AdobeAcrobatProDCContinuous","AdobeReaderDCContinuous",
    "Apache24SvrUnix","Apache24SvrWin","Apache24SiteUnix","Apache24SiteWin",
    "ApacheTomcatAS","CiscoXERtrNDM","CiscoXESwtchL2S","CiscoXESwtchNDM","Chrome",
    "IIS10Server","IIS10Site","IE11","JBoss","DotNET4","MSAccess2016","MSDefender",
    "MSEdge","MSExcel2016","MSExchange2016EdgeTP","MSExchange2016MB","MSExchange2019Edge",
    "MSExchange2019MB","MSGroove2013","MSOffice365","MSOfficeSystem2016","MSOneDrive",
    "MSOneNote2013","MSOneNote2016","MSOutlook2013","MSOutlook2016","MSPowerPoint2016",
    "MSProject2016","MSPublisher2016","MSSPDesigner2013","MSSkype2016","SQL2016DB",
    "SQL2016Instance","MSVisio2016","MSWord2016","Firefox","Oracle7","Oracle8","RGSRKE2",
    "RHEL8","RHEL9","TrellixENS10xLocal","Ubuntu20","Ubuntu22","Win10","Win11",
    "WinFirewall","WinServer2016","WinServer2019","WinServer2022","WinServerDNS"
]

# we will extract timestamp from filename via this regex:
TIMESTAMP_RX = re.compile(r'(\d{8}-\d{6})\.cklb$', re.I)
# prefix in JSON "title" field:
TITLE_PREFIX = "Evaluate-STIG_"

# AI Configuration
AI_CONFIG_FILE = "ai_config.json"

# ────────────────────────── HELPERS ───────────────────────────────────────────
def ts_now(fmt="%Y-%m-%d %H:%M:%S") -> str:
    return datetime.datetime.now().strftime(fmt)

def prompt_path(msg: str, default: str | None = None) -> str:
    inp = input(f"{msg}{f' [{default}]' if default else ''}: ").strip().strip('"').strip("'")
    return os.path.expanduser(inp or (default or ""))

def yes(msg: str, default=True) -> bool:
    yn = ("Y/n","y/N")[not default]
    while True:
        a = input(f"{msg} ({yn}) ").strip().lower()
        if not a: return default
        if a in ("y","yes"): return True
        if a in ("n","no"):  return False
        print("Enter y or n.")

def pretty(n: int) -> str:
    return f"{n:,}"

# ────────────────────────── AI FUNCTIONS ─────────────────────────────────────
def load_ai_config() -> Dict:
    """Load AI configuration"""
    default = {
        "base_url": "http://localhost:11434",
        "model": "codellama",
        "timeout": 120,
        "output_dir": "./generated_powershell"
    }
    try:
        if os.path.exists(AI_CONFIG_FILE):
            with open(AI_CONFIG_FILE, 'r') as f:
                config = json.load(f)
                # Merge with defaults
                for key, value in default.items():
                    if key not in config:
                        config[key] = value
                return config
        return default
    except Exception:
        return default

def save_ai_config(config: Dict):
    """Save AI configuration"""
    try:
        with open(AI_CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=2)
        print("✅ AI configuration saved.")
    except Exception as e:
        print(f"❌ Failed to save config: {e}")

def configure_ai():
    """Configure AI settings"""
    print("\n🤖 AI Configuration")
    print("=" * 40)
    
    config = load_ai_config()
    print(f"Current settings:")
    print(f"  Base URL: {config['base_url']}")
    print(f"  Model: {config['model']}")
    print(f"  Timeout: {config['timeout']}s")
    print(f"  Output Directory: {config['output_dir']}")
    
    if not yes("\nUpdate configuration?", False):
        return
    
    config['base_url'] = input(f"AI Base URL [{config['base_url']}]: ").strip() or config['base_url']
    config['model'] = input(f"Model name [{config['model']}]: ").strip() or config['model']
    
    timeout = input(f"Timeout seconds [{config['timeout']}]: ").strip()
    if timeout.isdigit():
        config['timeout'] = int(timeout)
    
    config['output_dir'] = input(f"Output directory [{config['output_dir']}]: ").strip() or config['output_dir']
    Path(config['output_dir']).mkdir(parents=True, exist_ok=True)
    
    # Test connection
    if test_ai_connection(config['base_url'], config['model']):
        save_ai_config(config)

def test_ai_connection(base_url: str, model: str) -> bool:
    """Test AI connection using OpenAI-compatible API only"""
    print(f"🔌 Testing OpenAI-compatible API at {base_url} with model '{model}'...")
    
    try:
        test_payload = {
            "model": model,
            "messages": [{"role": "user", "content": "Hello, test message"}],
            "max_tokens": 50
        }
        
        headers = {"Content-Type": "application/json"}
        
        response = requests.post(
            f"{base_url.rstrip('/')}/v1/chat/completions",
            json=test_payload,
            headers=headers,
            timeout=10
        )
        
        print(f"🔍 Response Status: {response.status_code}")
        
        if response.status_code == 200:
            try:
                result = response.json()
                if 'choices' in result and len(result['choices']) > 0:
                    message_content = result['choices'][0].get('message', {}).get('content', '')
                    print(f"✅ Connection successful! Model responded: '{message_content.strip()}'")
                    return True
                else:
                    print(f"❌ Unexpected response structure: {result}")
                    return False
            except Exception as je:
                print(f"❌ Failed to parse JSON response: {je}")
                return False
        else:
            print(f"❌ Request failed with status {response.status_code}")
            print(f"Response: {response.text[:300]}")
            return False
            
    except requests.exceptions.Timeout:
        print(f"❌ Request timed out after 10 seconds")
        return False
    except requests.exceptions.ConnectionError:
        print(f"❌ Cannot connect to {base_url}")
        return False
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        return False



def generate_powershell_code(stig_name: str, v_number: str, rule_title: str, description: str) -> str:
    """Generate PowerShell validation code using AI"""
    config = load_ai_config()
    
    prompt = f"""Generate a PowerShell validation script for this STIG requirement:

STIG: {stig_name}
V-Number: {v_number}
Rule: {rule_title}
Description: {description}

The script must follow this exact structure:
$ValidationResults = [PSCustomObject]@{{
    Results = ""
    Valid   = $true
}}
try {{
    # Your validation logic here
}} catch {{
    $ValidationResults.Results += "Error: $($_.Exception.Message)`n"
    $ValidationResults.Valid = $false
}}
return $ValidationResults

Requirements:
- Use proper error handling
- Include descriptive success (✅) and failure (❌) messages  
- Set Valid=$false for non-compliance
- Test the actual STIG requirement
- Use appropriate PowerShell cmdlets

Generate ONLY the PowerShell code, no markdown blocks or explanation."""
    
    try:
        # Use OpenAI-compatible format directly (matching your working PowerShell test)
        payload = {
            "model": config['model'],
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1,
            "max_tokens": 2000
        }
        
        headers = {"Content-Type": "application/json"}
        
        print(f"🤖 Sending request to {config['base_url']} with model {config['model']}...")
        
        response = requests.post(
            f"{config['base_url'].rstrip('/')}/v1/chat/completions",
            json=payload,
            headers=headers,
            timeout=config['timeout']
        )
        
        print(f"📡 Response status: {response.status_code}")
        
        if response.status_code == 200:
            try:
                result = response.json()
                content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
                if content.strip():
                    print("✅ Code generation successful!")
                    return clean_powershell_code(content)
                else:
                    return "# Error: Empty response from AI model"
            except Exception as parse_error:
                print(f"❌ Failed to parse response: {parse_error}")
                return f"# Error: Failed to parse AI response - {parse_error}"
        else:
            error_text = response.text[:500] if response.text else "No error details"
            print(f"❌ AI request failed: {response.status_code} - {error_text}")
            return f"# Error: AI request failed with status {response.status_code}\n# {error_text}"
            
    except requests.exceptions.Timeout:
        print(f"❌ Request timed out after {config['timeout']} seconds")
        return f"# Error: Request timed out after {config['timeout']} seconds"
    except requests.exceptions.ConnectionError:
        print(f"❌ Cannot connect to {config['base_url']}")
        return f"# Error: Cannot connect to {config['base_url']}"
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        return f"# Error: {str(e)}"

def clean_powershell_code(raw_code: str) -> str:
    """Clean and format PowerShell code"""
    # Remove markdown code blocks
    code = re.sub(r'^```(?:powershell)?\s*\n?', '', raw_code, flags=re.MULTILINE)
    code = re.sub(r'^```\s*$', '', code, flags=re.MULTILINE)
    return code.strip()

def parse_stig_library():
    """Parse STIG library and generate PowerShell code"""
    print("\n📚 STIG Library Parser & PowerShell Generator")
    print("=" * 50)
    
    library_path = Path(prompt_path("STIG Library directory path")).expanduser()
    if not library_path.exists():
        print("❌ Directory not found")
        return
    
    # Find ZIP files
    zip_files = list(library_path.glob("*.zip"))
    if not zip_files:
        print("❌ No ZIP files found")
        return
    
    print(f"📦 Found {len(zip_files)} ZIP files")
    
    # Parse STIGs
    stigs = {}
    for zip_file in zip_files:
        stig_info = parse_stig_zip(zip_file)
        if stig_info:
            stigs[stig_info['id']] = stig_info
            print(f"  ✅ {stig_info['title'][:60]}...")
    
    if not stigs:
        print("❌ No valid STIGs found")
        return
    
    # Select STIG
    print(f"\n📋 Available STIGs:")
    stig_list = list(stigs.values())
    for i, stig in enumerate(stig_list, 1):
        print(f"{i:2d}. {stig['title']}")
    
    try:
        choice = int(input("\nSelect STIG number: ")) - 1
        if not 0 <= choice < len(stig_list):
            print("❌ Invalid selection")
            return
        selected_stig = stig_list[choice]
    except ValueError:
        print("❌ Invalid input")
        return
    
    # Select rule
    rules = list(selected_stig['rules'].items())
    rules.sort(key=lambda x: x[0])  # Sort by V-number
    
    print(f"\n📋 Rules in {selected_stig['title']}:")
    for i, (v_num, rule_info) in enumerate(rules, 1):
        print(f"{i:3d}. {v_num} - {rule_info['title'][:70]}...")
    
    try:
        choice = int(input("\nSelect rule number: ")) - 1
        if not 0 <= choice < len(rules):
            print("❌ Invalid selection")
            return
        v_number, rule_info = rules[choice]
    except ValueError:
        print("❌ Invalid input")
        return
    
    # Generate PowerShell code
    print(f"\n🔄 Generating PowerShell for {v_number}...")
    code = generate_powershell_code(
        selected_stig['title'],
        v_number,
        rule_info['title'],
        rule_info['description']
    )
    
    # Display code
    print("\n" + "="*80)
    print("GENERATED POWERSHELL CODE")
    print("="*80)
    print(code)
    print("="*80)
    
    # Save option
    if yes("\nSave to file?", True):
        config = load_ai_config()
        output_dir = Path(config['output_dir'])
        output_dir.mkdir(parents=True, exist_ok=True)
        
        safe_name = re.sub(r'[^\w\s-]', '', selected_stig['title']).strip()[:30]
        safe_name = re.sub(r'[-\s]+', '_', safe_name)
        filename = f"{safe_name}_{v_number.replace('-', '_')}_{datetime.datetime.now():%Y%m%d_%H%M%S}"
        
        ext = input("File extension (ps1/txt) [ps1]: ").strip().lower() or "ps1"
        if ext not in ['ps1', 'txt']:
            ext = 'ps1'
        
        file_path = output_dir / f"{filename}.{ext}"
        
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                header = f"""# PowerShell STIG Validation Script
# Generated: {ts_now()}
# STIG: {selected_stig['title']}
# Rule: {v_number} - {rule_info['title']}
# Severity: {rule_info.get('severity', 'Unknown').upper()}
#
# Description:
# {rule_info['description'][:200]}{'...' if len(rule_info['description']) > 200 else ''}
#
# ============================================================================

"""
                f.write(header + code)
            
            print(f"✅ Saved to: {file_path}")
        except Exception as e:
            print(f"❌ Failed to save: {e}")

def parse_stig_zip(zip_path: Path) -> Optional[Dict]:
    """Parse a STIG ZIP file to extract rule information"""
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            # Look for XCCDF files
            xccdf_files = [f for f in zf.namelist() if f.endswith('-xccdf.xml')]
            if not xccdf_files:
                return None
            
            with zf.open(xccdf_files[0]) as f:
                return parse_xccdf_content(f.read(), zip_path.stem)
    except Exception:
        return None

def parse_xccdf_content(xml_content: bytes, zip_name: str) -> Optional[Dict]:
    """Parse XCCDF XML content"""
    try:
        root = ET.fromstring(xml_content)
        
        stig_info = {
            'id': zip_name,
            'title': '',
            'version': '',
            'rules': {}
        }
        
        # Extract title
        title_elem = root.find('.//{http://checklists.nist.gov/xccdf/1.1}title')
        if title_elem is not None and title_elem.text:
            stig_info['title'] = title_elem.text.strip()
        
        # Extract version
        version_elem = root.find('.//{http://checklists.nist.gov/xccdf/1.1}version')
        if version_elem is not None and version_elem.text:
            stig_info['version'] = version_elem.text.strip()
        
        # Extract rules (V-numbers)
        groups = root.findall('.//{http://checklists.nist.gov/xccdf/1.1}Group')
        
        for group in groups:
            group_id = group.get('id', '')
            if group_id.startswith('V-'):
                rule = group.find('.//{http://checklists.nist.gov/xccdf/1.1}Rule')
                if rule is not None:
                    rule_title_elem = rule.find('.//{http://checklists.nist.gov/xccdf/1.1}title')
                    rule_desc_elem = rule.find('.//{http://checklists.nist.gov/xccdf/1.1}description')
                    
                    rule_title = rule_title_elem.text.strip() if rule_title_elem is not None and rule_title_elem.text else ''
                    rule_desc = rule_desc_elem.text.strip() if rule_desc_elem is not None and rule_desc_elem.text else ''
                    
                    severity = rule.get('severity', 'medium')
                    
                    stig_info['rules'][group_id] = {
                        'title': rule_title,
                        'description': rule_desc,
                        'severity': severity
                    }
        
        return stig_info if stig_info['rules'] else None
        
    except Exception:
        return None

def generate_powershell():
    """Manual PowerShell generation"""
    print("\n🤖 Manual PowerShell Generator")
    print("=" * 40)
    
    stig_name = input("STIG Name: ").strip()
    if not stig_name:
        print("❌ STIG name required")
        return
        
    v_number = input("V-Number: ").strip()
    if not v_number:
        print("❌ V-Number required")
        return
        
    rule_title = input("Rule Title: ").strip()
    description = input("Description: ").strip()
    
    print(f"\n🔄 Generating PowerShell for {v_number}...")
    code = generate_powershell_code(stig_name, v_number, rule_title, description)
    
    print("\n" + "="*80)
    print("GENERATED POWERSHELL CODE")
    print("="*80)
    print(code)
    print("="*80)
    
    if yes("\nSave to file?", True):
        config = load_ai_config()
        output_dir = Path(config['output_dir'])
        output_dir.mkdir(parents=True, exist_ok=True)
        
        safe_stig = re.sub(r'[^\w\s-]', '', stig_name)[:30].replace(' ', '_')
        filename = f"{safe_stig}_{v_number.replace('-', '_')}_{datetime.datetime.now():%Y%m%d_%H%M%S}"
        
        ext = input("File extension (ps1/txt) [ps1]: ").strip().lower() or "ps1"
        file_path = output_dir / f"{filename}.{ext}"
        
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                header = f"""# PowerShell STIG Validation Script
# Generated: {ts_now()}
# STIG: {stig_name}
# Rule: {v_number} - {rule_title}
# ============================================================================

"""
                f.write(header + code)
            print(f"✅ Saved to: {file_path}")
        except Exception as e:
            print(f"❌ Failed to save: {e}")

# ═════════════════════════════════════════════════════════════════════════════
# EXISTING FUNCTIONS 
# ═════════════════════════════════════════════════════════════════════════════
def init_workbook():
    src = prompt_path("Template workbook path (.xlsx)")
    dst = Path(prompt_path("New workbook to create")).expanduser()
    if dst.exists() and not yes("Overwrite existing?", False):
        print("Aborted."); return
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(src, dst)
    wb = load_workbook(dst)
    headers = [c.value for c in wb["Chrome"][1]]
    for sn in ACTIVE_SHORTNAMES:
        if sn not in wb.sheetnames:
            ws = wb.create_sheet(sn)
            for i, h in enumerate(headers,1):
                ws.cell(1,i,h)
    for ws in wb.worksheets:
        if "AnswerKey Name" not in [c.value for c in ws[1]]:
            ws.cell(1, ws.max_column+1, "AnswerKey Name")
    wb.save(dst)
    print(f"✅ Created workbook at {dst}")

def clear_excel():
    path = Path(prompt_path("Workbook to clear (.xlsx)")).expanduser()
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for c in row: c.value = None
    wb.save(path)
    print("✅ Cleared all data rows.")

def newest_cklb(root: str) -> dict[str,str]:
    latest: dict[str,tuple[str,str]] = {}
    for dirpath, _, files in os.walk(root):
        for fn in files:
            if not fn.lower().endswith(".cklb"):
                continue
            m_ts = TIMESTAMP_RX.search(fn)
            if not m_ts:
                continue
            ts = m_ts.group(1)
            full = os.path.join(dirpath, fn)
            try:
                data = json.load(open(full, encoding="utf-8"))
            except Exception:
                continue
            title = data.get("title","")
            if not title.startswith(TITLE_PREFIX):
                continue
            short = title[len(TITLE_PREFIX):]
            prev = latest.get(short)
            if not prev or ts > prev[0]:
                latest[short] = (ts, full)
    return {short: path for short,(ts,path) in latest.items()}

def not_reviewed(path: str) -> list[str]:
    try:
        data = json.load(open(path, encoding="utf-8"))
    except Exception as e:
        print(f"⚠️ Skipping {path}: {e}")
        return []
    out = []
    for s in data.get("stigs", []):
        for r in s.get("rules", []):
            if str(r.get("status","")).lower() == "not_reviewed":
                vk = r.get("group_id") or r.get("vuln_id")
                if vk:
                    out.append(vk.strip())
    return out

def header_map(ws):
    hdrs = {str(c.value).strip().lower(): idx for idx,c in enumerate(ws[1],1) if c.value}
    def col(name: str):
        if name not in hdrs:
            idx = ws.max_column + 1
            ws.cell(1, idx, name)
            hdrs[name] = idx
        return hdrs[name]
    return {
        "vuln id":          col("vuln id"),
        "expectedstatus":   col("expectedstatus"),
        "validtruestatus":  col("validtruestatus"),
        "validtruecomment": col("validtruecomment"),
    }

def remove_old_comments(root: ET.Element):
    for e in list(root):
        if e.tag is ET.Comment and e.text and e.text.startswith("Script ran on"):
            root.remove(e)

def update_excel():
    wb_path  = Path(prompt_path("Workbook to update (.xlsx)")).expanduser()
    scan_dir = prompt_path("Directory containing .cklb files")
    vt_cmt   = input("ValidTrueComment (blank → STIG COMPLIANT): ").strip() or "STIG COMPLIANT"
    do_xml   = yes("Also update/create XML answer-files?", True)
    xml_dir  = Path(prompt_path("XML output directory")) if do_xml else None
    if xml_dir:
        xml_dir.mkdir(parents=True, exist_ok=True)

    stig_keys = {s: not_reviewed(p) for s,p in newest_cklb(scan_dir).items()}
    stig_keys = {s: ks for s,ks in stig_keys.items() if ks}

    wb = load_workbook(wb_path)
    for stig, keys in stig_keys.items():
        if stig not in wb.sheetnames:
            print(f"❌ Sheet '{stig}' missing – skipped.")
            continue
        ws = wb[stig]
        cols = header_map(ws)
        existing = {
            str(r[0]).strip()
            for r in ws.iter_rows(min_row=2,
                                  min_col=cols["vuln id"],
                                  max_col=cols["vuln id"],
                                  values_only=True)
            if r[0]
        }
        row = ws.max_row + 1
        for vk in keys:
            if vk in existing:
                continue
            ws.cell(row, cols["vuln id"], vk)
            ws.cell(row, cols["expectedstatus"], "Not_Reviewed")
            ws.cell(row, cols["validtruestatus"], "NotAFinding")
            ws.cell(row, cols["validtruecomment"], vt_cmt)
            row += 1
    wb.save(wb_path)
    print("✔️ Excel updated.")

    if not do_xml:
        return

    for stig, keys in stig_keys.items():
        xml = xml_dir / f"{stig}.xml"
        if xml.exists():
            tree = ET.parse(xml); root = tree.getroot()
        else:
            root = ET.Element("STIGComments", Name=stig)
            tree = ET.ElementTree(root)

        existing = {v.get("ID") for v in root.findall("Vuln")}
        added = []
        for vk in keys:
            if vk in existing:
                continue
            vul = ET.SubElement(root, "Vuln", ID=vk)
            ak  = ET.SubElement(vul, "AnswerKey", Name="DEFAULT")
            ET.SubElement(ak, "ExpectedStatus").text   = "Not_Reviewed"
            ET.SubElement(ak, "ValidationCode")
            ET.SubElement(ak, "ValidTrueStatus").text  = "NotAFinding"
            ET.SubElement(ak, "ValidTrueComment").text = vt_cmt
            ET.SubElement(ak, "ValidFalseStatus")
            ET.SubElement(ak, "ValidFalseComment")
            added.append(vk)

        if added:
            remove_old_comments(root)
            all_keys = sorted({v.get("ID") for v in root.findall("Vuln")})
            comment  = f"Script ran on {ts_now()} – Added V-keys: {', '.join(added)}"
            root.append(ET.Comment(comment))
            tree.write(xml, encoding="utf-8", xml_declaration=True)
            print(f"📝 {stig}: Added V-keys → {', '.join(added)}")

def generate_xml():
    wb_file = Path(prompt_path("Workbook path (.xlsx)")).expanduser()
    out_dir = Path(prompt_path("XML output directory")).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)

    xls = pd.ExcelFile(wb_file)
    for sheet in xls.sheet_names:
        df = xls.parse(sheet)
        xml = out_dir / f"{sheet}.xml"
        if xml.exists():
            tree = ET.parse(xml); root = tree.getroot()
        else:
            root = ET.Element("STIGComments", Name=sheet)
            tree = ET.ElementTree(root)

        added = []
        for _, row in df.iterrows():
            vid = row.get("Vuln ID", row.get("V Key", ""))
            vid = "" if pd.isna(vid) else str(vid).strip()
            if not vid:
                continue

            key = row.get("AnswerKey Name", "")
            key = "" if pd.isna(key) else str(key).strip()
            if not key:
                key = "DEFAULT"

            exp = row.get("ExpectedStatus", "")
            exp = "" if pd.isna(exp) else str(exp).strip() or "Not_Reviewed"
            vts = row.get("ValidTrueStatus", "")
            vts = "" if pd.isna(vts) else str(vts).strip() or "NotAFinding"
            vtc = row.get("ValidTrueComment", "")
            vtc = "" if pd.isna(vtc) else str(vtc).strip()

            vuln = next((v for v in root.findall("Vuln") if v.get("ID")==vid), None)
            if vuln is None:
                vuln = ET.SubElement(root, "Vuln", ID=vid)

            if any(ak.get("Name")==key for ak in vuln.findall("AnswerKey")):
                continue

            ak = ET.SubElement(vuln, "AnswerKey", Name=key)
            ET.SubElement(ak, "ExpectedStatus").text   = exp
            ET.SubElement(ak, "ValidationCode")
            ET.SubElement(ak, "ValidTrueStatus").text  = vts
            ET.SubElement(ak, "ValidTrueComment").text = vtc
            ET.SubElement(ak, "ValidFalseStatus")
            ET.SubElement(ak, "ValidFalseComment")

            added.append(f"{vid}({key})")

        if added:
            comment = f"Script ran on {ts_now()} – Added: {', '.join(added)}"
            root.append(ET.Comment(comment))
            tree.write(xml, encoding="utf-8", xml_declaration=True)
            print(f"🛈 {sheet}: Added → {', '.join(added)}")
        else:
            print(f"✓ {sheet}: no new entries")

def extract_cklb(path: str):
    try:
        data = json.load(open(path, encoding="utf-8"))
    except Exception as e:
        print(f"❌ {path}: {e}")
        return None, [], 0, Counter()

    host = (
        data.get("host_name")
        or (data.get("targets",[{}])[0].get("host_name") if isinstance(data.get("targets"), list) else None)
        or data.get("target_data",{}).get("host_name")
        or "Unknown Host"
    )

    findings = []
    total = 0
    counts = Counter()
    for s in data.get("stigs", []):
        nm = s.get("stig_name","Unknown STIG")
        sc = Counter(r.get("status","").strip().lower() for r in s.get("rules", []))
        counts.update(sc)
        n = sum(sc.values())
        total += n
        findings.append((nm, sc))

    return host, findings, total, counts

def report():
    root_dir = Path(prompt_path("Directory to scan recursively for .cklb")).expanduser()
    files = list(root_dir.rglob("*.cklb"))
    if not files:
        print("❌ No .cklb files found under that directory."); return

    out = Path(prompt_path("Output Markdown file or folder")).expanduser()
    if out.is_dir():
        out = out / f"stig_report_{datetime.datetime.now():%Y%m%d-%H%M%S}.md"

    summary = defaultdict(int)
    details = []
    grand = Counter()
    for fp in files:
        host, findings, tot, cnt = extract_cklb(str(fp))
        if host:
            summary[host] += tot
            details.append((fp, host, findings))
            grand.update(cnt)

    with open(out, "w", encoding="utf-8") as fh:
        fh.write("# STIG Open Findings Summary Report\n\n")
        fh.write("## Summary by Host\n\n| Host | Total Findings |\n|------|----------------|\n")
        for h,t in summary.items():
            fh.write(f"| {h} | {pretty(t)} |\n")

        fh.write("\n## Detailed Findings by File\n\n")
        for fp, host, findings in details:
            fh.write(f"### File: `{fp.name}`\n- Host: **{host}**\n")
            for nm, sc in findings:
                cnt = sum(sc.values())
                fh.write(f"  - STIG: *{nm}* — **{pretty(cnt)}** findings\n")
                for st,c in sorted(sc.items()):
                    fh.write(f"    - {st.replace('_',' ').title()}: {pretty(c)}\n")
            fh.write("\n")

        impl = grand.get("not_a_finding",0)
        op   = grand.get("open",0)
        tot  = impl + op
        pct  = (impl/tot*100) if tot else 0
        fh.write("## STIG Implementation Summary\n\n")
        fh.write(f"- Total Evaluated: **{pretty(tot)}**\n")
        fh.write(f"- Compliant (Not a Finding): **{pretty(impl)}**\n")
        fh.write(f"- Non-compliant (Open): **{pretty(op)}**\n")
        fh.write(f"\n**Overall Implementation: {pct:.2f}%**\n")
        fh.write(f"\n---\n_Report generated on {ts_now()}_\n")

    print(f"✅ Markdown report saved to {out}")

# ═════════════════════════════════════════════════════════════════════════════
# 6. INTERACTIVE MENU (Updated)
# ═════════════════════════════════════════════════════════════════════════════
def menu():
    while True:
        print("\n" + "="*60)
        print("ESTIG Tool - Enhanced with AI PowerShell Generation")
        print("="*60)
        print("1) Init workbook       2) Clear         3) Update")
        print("4) Generate XML        5) Report        6) Generate PowerShell")
        print("7) STIG Library        8) AI Config     0) Exit")
        print("="*60)
        
        c = input("> ").strip()
        if c == "0":
            break
        {
            "1": init_workbook,
            "2": clear_excel,
            "3": update_excel,
            "4": generate_xml,
            "5": report,
            "6": generate_powershell,
            "7": parse_stig_library,
            "8": configure_ai
        }.get(c, lambda: print("Invalid choice"))()

# ═════════════════════════════════════════════════════════════════════════════
# 7. CLI ENTRYPOINT (Updated)
# ═════════════════════════════════════════════════════════════════════════════
EPILOG = """
EXAMPLES
  python estig_tool.py -u
  python estig_tool.py -c -u -g
  python estig_tool.py -r
  python estig_tool.py -p
  python estig_tool.py -l
"""

def main():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawTextHelpFormatter,
        description="All-in-one STIG Excel/XML/report helper with AI PowerShell generation",
        epilog=EPILOG
    )
    parser.add_argument("-i","--init",        action="store_true",
                        help="Create a workbook from a template.")
    parser.add_argument("-c","--clear",       action="store_true",
                        help="Clear all data rows in a workbook.")
    parser.add_argument("-u","--update",      action="store_true",
                        help="Import Not_Reviewed V-keys into the workbook and optionally XML.")
    parser.add_argument("-m","--manualestig", action="store_true",
                        help="Alias for --update (legacy).")
    parser.add_argument("-g","--generate",    action="store_true",
                        help="Generate/refresh XML answer-files from the workbook.")
    parser.add_argument("-r","--report",      action="store_true",
                        help="Recursively scan for .cklb and build Markdown report.")
    parser.add_argument("-p","--powershell",  action="store_true",
                        help="Generate PowerShell validation code using AI.")
    parser.add_argument("-a","--aiconfig",    action="store_true",
                        help="Configure AI model settings.")
    parser.add_argument("-l","--library",     action="store_true",
                        help="Browse STIG library and generate PowerShell from rules.")
    
    args = parser.parse_args()

    steps = []
    if args.init:
        steps.append(init_workbook)
    if args.clear:
        steps.append(clear_excel)
    if args.update or args.manualestig:
        steps.append(update_excel)
    if args.generate:
        steps.append(generate_xml)
    if args.report:
        steps.append(report)
    if args.powershell:
        steps.append(generate_powershell)
    if args.aiconfig:
        steps.append(configure_ai)
    if args.library:
        steps.append(parse_stig_library)

    if not steps:
        menu()
    else:
        for fn in steps:
            try:
                fn()
            except KeyboardInterrupt:
                sys.exit("\nInterrupted by user.")
            except Exception as e:
                print(f"❌ Error: {e}")

if __name__ == "__main__":

    main()
